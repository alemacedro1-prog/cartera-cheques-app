from __future__ import annotations

import io
import re
import unicodedata
import zipfile
from datetime import date, datetime
from typing import BinaryIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ALLOWED_TYPES = ("CH24", "CH48", "CPD", "ECHEQ", "ECHEQDIF")
BANK_FILTER_OPTIONS = ("Macro", "Galicia", "Nación")
REJECTED_STATES = {"RE", "RC"}
PENDING_ACCREDITATION_STATES = {"PS"}
MAX_FILE_BYTES = 15 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 150 * 1024 * 1024
MAX_ROWS = 100_000


class ConcentradorError(ValueError):
    """Error de entrada seguro para mostrar al usuario."""


def _clean(value) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text in {"60", "nan", "None"} else text


def _identifier(value) -> str:
    text = _clean(value)
    if not text or text == "0":
        return ""
    return text[:-2] if re.fullmatch(r"-?\d+\.0", text) else text


def bank_filter_group(value) -> str:
    """Agrupa los bancos operativos admitiendo código BCRA o denominación."""
    text = _identifier(value)
    if not text:
        return ""

    numeric = re.fullmatch(r"0*(\d+)", text)
    if numeric:
        by_code = {"7": "Galicia", "11": "Nación", "285": "Macro"}
        if numeric.group(1) in by_code:
            return by_code[numeric.group(1)]

    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(character for character in normalized if not unicodedata.combining(character)).upper()
    if "MACRO" in normalized:
        return "Macro"
    if "GALICIA" in normalized:
        return "Galicia"
    if "NACION" in normalized:
        return "Nación"
    return ""


def _system_state(value) -> tuple[str, str]:
    """Devuelve el valor original y un código de estado normalizado."""
    original = _clean(value).upper()
    if not original:
        return "", ""
    patterns = {
        "PS": r"(?<![A-Z])P[\s./_-]*S(?![A-Z])",
        "RE": r"(?<![A-Z])R[\s./_-]*E(?![A-Z])",
        "RC": r"(?<![A-Z])R[\s./_-]*C(?![A-Z])",
        "AC": r"(?<![A-Z])A[\s./_-]*C(?![A-Z])",
    }
    for code, pattern in patterns.items():
        if re.search(pattern, original):
            return original, code
    return original, original


def _date_value(value):
    if value is None or pd.isna(value) or value in (0, 60, "0", "60"):
        return pd.NaT
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return pd.Timestamp(value).normalize()
    if isinstance(value, (int, float)) and 20_000 < value < 80_000:
        return pd.Timestamp("1899-12-30") + pd.to_timedelta(value, unit="D")
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    return parsed.normalize() if not pd.isna(parsed) else pd.NaT


def _first(row: pd.Series, *columns: str):
    for column in columns:
        if column in row.index and _clean(row[column]):
            return row[column]
    return ""


def _extract_observation_receipt(observation: str) -> str:
    patterns = (
        r"(?:recibo|comprobante|cbte\.?)(?:\s+(?:relacionado|relación))?\s*(?:n[roº°]*\.?)?\s*[:#-]?\s*(\d{2,}(?:[-/]\d{2,})*)",
        r"existe en (?:el\s+)?(?:resumen bancario|la recaudadora)\s*(?:n[roº°]*\.?)?\s*[:#-]?\s*(\d{2,}(?:[-/]\d{2,})*)",
    )
    for pattern in patterns:
        match = re.search(pattern, observation, re.I)
        if match:
            return match.group(1)
    internal_receipt = re.search(r"(?<!\d)(75\d+)(?!\d)", observation)
    if internal_receipt:
        return internal_receipt.group(1)
    receipt_like = re.search(r"(?<![\d-])(\d{2,6}-\d{2,10})(?![\d-])", observation)
    if receipt_like:
        return receipt_like.group(1)
    return ""


def validate_upload(file_bytes: bytes) -> None:
    if not file_bytes:
        raise ConcentradorError("El archivo está vacío.")
    if len(file_bytes) > MAX_FILE_BYTES:
        raise ConcentradorError("El archivo supera el límite permitido de 15 MB.")
    if not file_bytes.startswith(b"PK"):
        raise ConcentradorError("El archivo no es un Excel XLSX/XLSM válido.")
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            members = archive.infolist()
            if not members or "[Content_Types].xml" not in {item.filename for item in members}:
                raise ConcentradorError("El archivo no tiene una estructura Excel válida.")
            total = sum(item.file_size for item in members)
            if total > MAX_UNCOMPRESSED_BYTES:
                raise ConcentradorError("El Excel expandido supera el límite seguro de 150 MB.")
            for item in members:
                if item.compress_size and item.file_size / item.compress_size > 200:
                    raise ConcentradorError("El archivo fue rechazado por compresión anómala.")
    except zipfile.BadZipFile as exc:
        raise ConcentradorError("El archivo Excel está dañado o incompleto.") from exc


def _header_row(file_obj: BinaryIO) -> int:
    preview = pd.read_excel(file_obj, sheet_name=0, header=None, nrows=35, dtype=object)
    file_obj.seek(0)
    for idx, row in preview.iterrows():
        values = {_clean(value) for value in row.tolist()}
        if "MCR-Medio de pago" in values and "MCR-Importe instr." in values:
            return int(idx)
    raise ConcentradorError(
        "No encontré los encabezados esperados del CONRENPF. Verificá que sea el archivo original."
    )


def read_concentrador(file_bytes: bytes) -> pd.DataFrame:
    validate_upload(file_bytes)
    stream = io.BytesIO(file_bytes)
    header = _header_row(stream)
    raw = pd.read_excel(stream, sheet_name=0, header=header, dtype=object)
    if len(raw) > MAX_ROWS:
        raise ConcentradorError("El archivo supera el límite permitido de 100.000 filas.")
    raw.columns = [str(column).strip() for column in raw.columns]
    required = {"MCR-Medio de pago", "MCR-Importe instr.", "MCR-Estado instr."}
    missing = sorted(required.difference(raw.columns))
    if missing:
        raise ConcentradorError("Faltan columnas obligatorias: " + ", ".join(missing))
    raw["Fila fuente"] = raw.index + header + 2
    return raw


def build_portfolio(raw: pd.DataFrame, cutoff: date | pd.Timestamp | None = None) -> pd.DataFrame:
    required = {"MCR-Medio de pago", "MCR-Importe instr.", "MCR-Estado instr."}
    missing = sorted(required.difference(raw.columns))
    if missing:
        raise ConcentradorError("Faltan columnas obligatorias: " + ", ".join(missing))
    cutoff_ts = pd.Timestamp(cutoff or date.today()).normalize()
    target = raw[raw["MCR-Medio de pago"].map(lambda value: _clean(value).upper()).isin(ALLOWED_TYPES)].copy()
    if "Fila fuente" not in target:
        target["Fila fuente"] = target.index + 2
    records: list[dict] = []

    for number, (_, row) in enumerate(target.iterrows(), start=1):
        instrument_type = _clean(row.get("MCR-Medio de pago")).upper()
        observation = _clean(row.get("Observación"))
        cpb_relation = _identifier(row.get("Nro Cpb Relación"))
        observation_receipt = _extract_observation_receipt(observation)

        receipt, source = (
            (observation_receipt, "Observación") if observation_receipt else
            (cpb_relation, "Nro Cpb Relación") if cpb_relation else
            ("", "Sin vínculo detectado")
        )

        original_state_text, original_state = _system_state(_first(row, "MCR-Estado instr.", "Estado"))
        rejection_code = _clean(row.get("MCR-Código rechazo"))
        rejection_reason = _clean(row.get("MCR-Motivo rechazo"))
        entry_date = _date_value(_first(row, "MCR-Fecha pago", "Fecha Movimiento"))
        accreditation_date = _date_value(_first(row, "MCR-Fecha acredit.", "Fecha acreditación"))
        due_date = _date_value(row.get("MCR-Fecha vencim."))
        days_to_due = int((due_date - cutoff_ts).days) if not pd.isna(due_date) else pd.NA
        # El estado del sistema es la fuente de verdad. Código y motivo son
        # informativos y no convierten por sí solos un movimiento en rechazado.
        rejected = original_state in REJECTED_STATES
        if rejected:
            state = "Rechazado"
        elif original_state in PENDING_ACCREDITATION_STATES:
            state = "Pendiente de acreditación"
        elif not pd.isna(accreditation_date):
            state = "Acreditado" if accreditation_date <= cutoff_ts else "Pendiente de acreditación"
        elif original_state == "AC":
            state = "Acreditado"
        elif pd.isna(due_date):
            state = "Sin vencimiento"
        elif days_to_due < 0:
            state = "Vencido"
        elif days_to_due == 0:
            state = "Vence hoy"
        else:
            state = "Pendiente"

        if pd.isna(days_to_due): bucket = "Sin vencimiento"
        elif days_to_due < 0: bucket = "Vencidos"
        elif days_to_due == 0: bucket = "Hoy"
        elif days_to_due <= 7: bucket = "1-7 días"
        elif days_to_due <= 15: bucket = "8-15 días"
        elif days_to_due <= 30: bucket = "16-30 días"
        elif days_to_due <= 60: bucket = "31-60 días"
        else: bucket = ">60 días"

        alerts = []
        if rejected: alerts.append("RECHAZADO")
        if state == "Pendiente de acreditación": alerts.append("PENDIENTE DE ACREDITACIÓN")
        if not receipt: alerts.append("SIN RECIBO ASOCIADO")
        if state == "Vencido": alerts.append("VENCIDO")
        elif state == "Vence hoy": alerts.append("VENCE HOY")
        elif state == "Pendiente" and days_to_due <= 7: alerts.append("VENCE ≤ 7 DÍAS")
        amount = pd.to_numeric(_first(row, "MCR-Importe instr.", "Importe"), errors="coerce")

        records.append({
            "ID cartera": f"CHQ-{number:04d}", "Fila fuente": int(row["Fila fuente"]),
            "Tipo": instrument_type, "Cliente": _clean(_first(row, "MCR-Nombre cliente", "Nombre")),
            "CUIT cliente": _identifier(row.get("MCR-CUIT cliente")),
            "N° cliente": _identifier(row.get("MCR-NRO. de cliente")),
            "N° cheque / eCheq": _identifier(_first(row, "MCR-Número de cheque", "Nro de Cheque", "MCR-Nro instrumento")),
            "Banco cheque": _identifier(_first(row, "MCR-Banco", "Banco Del Cheque")),
            "Sucursal cheque": _identifier(_first(row, "MCR-Sucursal", "Suc Del Cheque")),
            "Cuenta cheque": _identifier(row.get("MCR-Cuenta de cheque")), "CUIT emisor": _identifier(row.get("MCR-CUIT emisor")),
            "Importe": float(amount) if not pd.isna(amount) else 0.0,
            "Fecha ingreso / pago": entry_date, "Fecha acreditación": accreditation_date,
            "Fecha vencimiento": due_date, "Días al vencimiento": days_to_due, "Tramo vencimiento": bucket,
            "Estado recibo": "Tomado" if receipt else "Sin recibo asociado",
            "Recibo relacionado": receipt, "Fuente del vínculo": source,
            "Estado calculado": state, "Código estado": original_state, "Estado original": original_state_text,
            "Código rechazo": rejection_code, "Motivo rechazo": rejection_reason, "Observaciones": observation,
            "Alertas": " · ".join(alerts), "Nro Cpb Relación": cpb_relation,
            "Banco depósito": _clean(_first(row, "Nombre del banco", "MCR-Banco depósito")),
            "MCR-ID pago": _identifier(row.get("MCR-ID pago")), "MCR-ID instrumento": _identifier(row.get("MCR-ID instrumento")),
            "N° operación": _identifier(row.get("MCR-Número operación")),
        })
    return pd.DataFrame.from_records(records)


def portfolio_from_bytes(file_bytes: bytes, cutoff: date | None = None):
    raw = read_concentrador(file_bytes)
    return build_portfolio(raw, cutoff), raw


def rejected_monthly_summary(portfolio: pd.DataFrame) -> pd.DataFrame:
    """Resume rechazos por mes y cliente usando la mejor fecha operativa disponible."""
    columns = ["Mes", "Cliente", "Cantidad de rechazados", "Importe rechazado", "Importe promedio"]
    if portfolio.empty or "Estado calculado" not in portfolio:
        return pd.DataFrame(columns=columns)

    rejected = portfolio[portfolio["Estado calculado"].eq("Rechazado")].copy()
    if rejected.empty:
        return pd.DataFrame(columns=columns)

    dates = pd.Series(pd.NaT, index=rejected.index, dtype="datetime64[ns]")
    for date_column in ("Fecha ingreso / pago", "Fecha vencimiento", "Fecha acreditación"):
        if date_column in rejected:
            dates = dates.fillna(pd.to_datetime(rejected[date_column], dayfirst=True, errors="coerce"))
    rejected["Mes"] = dates.dt.to_period("M").dt.to_timestamp()
    rejected = rejected[rejected["Mes"].notna()].copy()
    if rejected.empty:
        return pd.DataFrame(columns=columns)

    rejected["Cliente"] = rejected.get("Cliente", pd.Series(index=rejected.index, dtype=object)).fillna("").astype(str).str.strip().replace("", "Sin cliente")
    rejected["Importe"] = (
        pd.to_numeric(rejected["Importe"], errors="coerce").fillna(0)
        if "Importe" in rejected else 0.0
    )
    summary = rejected.groupby(["Mes", "Cliente"], as_index=False).agg(
        **{
            "Cantidad de rechazados": ("Importe", "size"),
            "Importe rechazado": ("Importe", "sum"),
            "Importe promedio": ("Importe", "mean"),
        }
    )
    return summary[columns].sort_values(["Mes", "Importe rechazado", "Cliente"], ascending=[True, False, True]).reset_index(drop=True)


def _excel_safe(frame: pd.DataFrame) -> pd.DataFrame:
    safe = frame.copy()
    for column in safe.select_dtypes(include=["object", "string"]).columns:
        safe[column] = safe[column].map(
            lambda value: "'" + value if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")) else value
        )
    return safe


def export_excel(portfolio: pd.DataFrame, raw: pd.DataFrame, cutoff: date) -> bytes:
    source_rows = set(portfolio.get("Fila fuente", pd.Series(dtype=int)).dropna().astype(int))
    scoped_raw = raw[raw["Fila fuente"].isin(source_rows)].copy() if "Fila fuente" in raw else raw.iloc[0:0]
    scoped_raw = scoped_raw.drop(columns=["MCR-Número de recibo", "Nro Cpb Relacionado"], errors="ignore")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        with_receipt = portfolio["Estado recibo"].eq("Tomado")
        summary = pd.DataFrame({
            "Indicador": [
                "Fecha de corte", "Instrumentos", "Importe total", "En cartera (con recibo)",
                "Cheques con recibo", "Importe con recibo", "Cheques sin recibo", "Importe sin recibo", "Rechazados",
            ],
            "Valor": [
                pd.Timestamp(cutoff), len(portfolio), portfolio["Importe"].sum(),
                portfolio.loc[with_receipt, "Importe"].sum(), int(with_receipt.sum()),
                portfolio.loc[with_receipt, "Importe"].sum(), int((~with_receipt).sum()),
                portfolio.loc[~with_receipt, "Importe"].sum(),
                portfolio.loc[portfolio["Estado calculado"].eq("Rechazado"), "Importe"].sum(),
            ],
        })
        summary.to_excel(writer, sheet_name="Resumen", index=False)
        _excel_safe(portfolio).to_excel(writer, sheet_name="Cartera", index=False)
        _excel_safe(scoped_raw).to_excel(writer, sheet_name="Datos fuente filtrados", index=False)
        for sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]; ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions; ws.sheet_view.showGridLines = False
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="17365D"); cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            for cells in ws.columns:
                letter = get_column_letter(cells[0].column); sample = [str(c.value or "") for c in list(cells)[:150]]
                ws.column_dimensions[letter].width = min(max(max(map(len, sample), default=8) + 2, 11), 35)
        ws = writer.book["Cartera"]
        for cell in ws[get_column_letter(list(portfolio.columns).index("Importe") + 1)][1:]: cell.number_format = '$#,##0.00;[Red]($#,##0.00);-'
        for name in ("Fecha ingreso / pago", "Fecha acreditación", "Fecha vencimiento"):
            for cell in ws[get_column_letter(list(portfolio.columns).index(name) + 1)][1:]: cell.number_format = "dd/mm/yyyy"
    return output.getvalue()


def format_currency(value: float) -> str:
    return "$ " + f"{value:,.0f}".replace(",", ".")
